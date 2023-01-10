
from openpyxl import load_workbook
import os
import csv
from collections import Counter
from random import shuffle



folder_path = "/Users/anahaza/Desktop/Accounts/Invensense/test-setdec2022" #folder with all your files. Teset sets or annotations
new_testset = os.path.join(folder_path,"testset_output.csv")
rows = []
file_with_eids_path = "/Users/anahaza/Desktop/Accounts/Invensense/barista-Intent-tdk.csv" #Download this file from your tenant. You need the intents one

eids_dictionary = {}
with open(file_with_eids_path) as f:
    f.readline()
    reader = csv.reader(f, delimiter=",", quotechar='"')
    for line in reader:
        eid = line[6]
        name = line[14]
        eids_dictionary[name] = eid

def process_excel_file(file): #for annotations
    wb = load_workbook(os.path.join(folder_path,file))
    if "Annotation_Sheet" not in wb.sheetnames:
        return
    ws = wb["Annotation_Sheet"]
    i = 0
    header = list(ws.rows)[0]
    for column in header:
        if column.value == "Interaction Text":
            phrase_number = i
        elif column.value == "Actual Matched Intent":
            intent_number = i
        elif column.value == "Annotation":
            annotation_number = i 
        i += 1
    for row in list(ws.rows)[1:]:
        annotation = str(row[annotation_number].value) or ""
        if annotation.lower() == "y":
            phrase = row[phrase_number]
            intent = row[intent_number]
            rows.append((str(phrase.value).strip(),intent.value,"TO_TRIAGE"))
        elif annotation.lower() != "invalid":
            phrase = row[phrase_number]
            rows.append((str(phrase.value).strip(),"","TO_TRIAGE"))


def process_csv_file(file): #from testset
    with open(os.path.join(folder_path, file)) as f:
        header = f.readline()
        reader = csv.reader(f, delimiter=",", quotechar='"')
        for line in reader:
            phrase = line[0]
            intent = line[2]
            status = line[1]
            rows.append((phrase.strip(), intent,status))


#Function definition to select a sample of the whole data set taking into account the rows that are most frequent
def sample(rows, sample_size):
    #rows = [row[0] for row in rows]
    
    #Order by frequency, divide in two sets, "most frequent" and "random"
    freqs = Counter(rows)
    rows_frequent = []
    rows_random = [] #I need a ticket 
    for row, freq in freqs.most_common():
        if freq > 100:
            rows_frequent.append((row, freq/len(rows)))
        else:
            rows_random.append((row, freq/len(rows)))

    #Add all "most frequent" rows, maintaining the same proportion
    sample = []
    for row, percentage in rows_frequent:
        how_many = int(sample_size * percentage)
        for i in range(0, how_many):
            if row not in sample:
                sample.append(row)

    #Add random rows as needed to get to the desired corpus size
    shuffle(rows_random)
    for row, _ in rows_random:
        if row not in sample:
            sample.append(row)
        if len(sample) == sample_size:
            break
            
    return sample

#just calling the funtions
for file in os.listdir(folder_path):
    if file.endswith("xlsx"):
        process_excel_file(file)
    elif file.endswith("csv") and file != "testset_output.csv":
        process_csv_file(file)

sampled = sample(rows,300)#update number of interactions you want here

#creating a testset
with open(new_testset, 'w', newline='') as csvfile:
    test_writer = csv.writer(csvfile, delimiter=',',
                            quotechar= '"', quoting=csv.QUOTE_MINIMAL)
    test_writer.writerow(["phrase","expected_intent_eid","status","source_intent_name"])
    for row in sampled: 
        intent_eid = eids_dictionary.get(row[1], "")
        test_writer.writerow([row[0],intent_eid,row[2],row[1]])